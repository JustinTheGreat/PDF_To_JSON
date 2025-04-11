import openpyxl
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from Components.json_processor import JsonProcessor
from Components.text_filters import TextFilter

class ExcelGenerator:
    def __init__(self):
        """Initialize Excel generator with default styling."""
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
            'border': self.thin_border
        }
        
        self.subtitle_style = {
            'font': Font(bold=True, italic=True),
            'fill': PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
            'border': self.thin_border
        }
    
    def apply_cell_style(self, cell, style_dict):
        """Apply a dictionary of styles to a cell."""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def create_excel_file(self, all_json_data, output_excel_path, filter_text="", 
                          apply_value_filters=True, callback=None):
        """
        Create an Excel file based on the JSON data.
        
        Args:
            all_json_data: Dictionary of JSON data keyed by filename
            output_excel_path: Path where the Excel file will be saved
            filter_text: Text to remove from filenames
            apply_value_filters: Whether to apply text filters to values
            callback: Function to call with status updates (optional)
        """
        try:
            # Debug helper function
            def debug(message):
                if callback:
                    callback("debug", f"Excel Generator: {message}")
                print(message)
            
            debug(f"Starting Excel generation with {len(all_json_data)} files")
            debug(f"Output path: {output_excel_path}")
            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            debug("Created new workbook")
            
            # Dictionary to track worksheets by title and their current row
            worksheets = {}  # {title: {'sheet': worksheet, 'next_row': row_number}}
            
            # First pass: analyze all reports to find the maximum list lengths
            debug("First pass: analyzing all reports to find maximum list lengths for each field")
            max_list_lengths = {}  # {title: {field_name: max_length}}
            
            for file_name, file_json_data in all_json_data.items():
                # Convert to list if not already
                reports_to_process = file_json_data if isinstance(file_json_data, list) else [file_json_data]
                
                for report in reports_to_process:
                    # Extract the title
                    title = None
                    if isinstance(report, dict):
                        title = report.get('title', None)
                    
                    if title is None:
                        title = f"Report_{file_name}"
                    
                    # Make a safe title for Excel worksheet
                    safe_title = ''.join(c for c in title if c not in '\\/:*?[]')
                    safe_title = safe_title[:31]  # Excel sheet name limit
                    
                    # Extract fields from the report
                    fields = {}
                    if isinstance(report, dict) and 'fields' in report:
                        fields = report.get('fields', {})
                    elif isinstance(report, dict):
                        fields = report
                    
                    # Initialize max_list_lengths for this title if not already done
                    if safe_title not in max_list_lengths:
                        max_list_lengths[safe_title] = {}
                    
                    # Check each field for list lengths
                    for key, value in fields.items():
                        if isinstance(value, list):
                            current_max = max_list_lengths[safe_title].get(key, 0)
                            max_list_lengths[safe_title][key] = max(current_max, len(value))
            
            # Print max list lengths for debugging
            for title, field_lengths in max_list_lengths.items():
                debug(f"Maximum list lengths for worksheet '{title}':")
                for field, length in field_lengths.items():
                    if length > 1:
                        debug(f"  Field '{field}': {length} items")
            
            # Process each file's JSON data
            total_files = len(all_json_data)
            total_reports_processed = 0
            
            for file_index, (file_name, file_json_data) in enumerate(all_json_data.items()):
                # Update progress if callback provided
                if callback:
                    callback("status", f"Processing {file_name}...")
                    callback("progress", (file_index + 1) / total_files * 100)
                
                debug(f"Processing file {file_index+1}/{total_files}: {file_name}")
                
                # Check the type of the JSON data
                if isinstance(file_json_data, dict):
                    debug(f"  File data is a dictionary with {len(file_json_data)} keys")
                    if file_json_data:
                        debug(f"  Keys: {list(file_json_data.keys())[:5]}...")
                    # Check if there's a title key which suggests it's a single report
                    if 'title' in file_json_data:
                        debug("  Found 'title' key, treating as a single report")
                        reports_to_process = [file_json_data]
                    else:
                        debug("  No 'title' key found, treating entire dictionary as a single report")
                        reports_to_process = [file_json_data]
                elif isinstance(file_json_data, list):
                    debug(f"  File data is a list with {len(file_json_data)} items")
                    if file_json_data and isinstance(file_json_data[0], dict):
                        sample_keys = list(file_json_data[0].keys())[:5]
                        debug(f"  First item keys: {sample_keys}")
                    reports_to_process = file_json_data
                else:
                    debug(f"  File data is a {type(file_json_data).__name__}, not a dict or list")
                    debug("  Wrapping in a list for processing")
                    reports_to_process = [file_json_data]
                
                debug(f"  Will process {len(reports_to_process)} reports from this file")
                
                # Process each report in the JSON data
                for report_index, report in enumerate(reports_to_process):
                    debug(f"  Processing report {report_index+1}/{len(reports_to_process)}")
                    
                    # Extract the title
                    title = None
                    if isinstance(report, dict):
                        title = report.get('title', None)
                    
                    if title is None:
                        title = f"Report_{file_name}_{report_index}"
                        debug(f"  No title found, using generated title: {title}")
                    else:
                        debug(f"  Report title: {title}")
                    
                    # Process this report
                    total_reports_processed += 1
                    
                    # Make a safe title for Excel worksheet
                    safe_title = ''.join(c for c in title if c not in '\\/:*?[]')
                    safe_title = safe_title[:31]  # Excel sheet name limit
                    
                    # Check if we already have a worksheet for this title
                    if safe_title in worksheets:
                        debug(f"  Adding to existing worksheet: {safe_title}")
                        worksheet = worksheets[safe_title]['sheet']
                        next_row = worksheets[safe_title]['next_row']
                    else:
                        # Create a new worksheet
                        worksheet = workbook.create_sheet(title=safe_title)
                        debug(f"  Created new worksheet: {safe_title}")
                        
                        # Analyze the structure to set up headers, but augment with max list lengths
                        debug("  Analyzing report structure")
                        structure_info = JsonProcessor.analyze_json_structure([report], False)
                        
                        # Override nesting_depth with the maximum values we found in the first pass
                        if safe_title in max_list_lengths:
                            for key, max_length in max_list_lengths[safe_title].items():
                                if max_length > 1:
                                    structure_info['keys'].add(key)
                                    structure_info['nesting_depth'][key] = max_length
                                    structure_info['needs_subtitles'] = True
                        
                        debug(f"  Found {len(structure_info['keys'])} unique keys")
                        debug(f"  Needs subtitles: {structure_info['needs_subtitles']}")
                        
                        # Set up the headers
                        self.setup_headers(worksheet, structure_info)
                        
                        # Start with row after headers
                        if structure_info['needs_subtitles']:
                            next_row = 3  # Start at row 3 (after header and subtitle rows)
                        else:
                            next_row = 2  # Start at row 2 (after header row)
                        
                        # Store worksheet info
                        worksheets[safe_title] = {
                            'sheet': worksheet,
                            'next_row': next_row,
                            'structure_info': structure_info,
                            'column_count': self.get_column_count(structure_info)
                        }
                    
                    # Extract fields from the report
                    if isinstance(report, dict) and 'fields' in report:
                        fields = report.get('fields', {})
                        debug(f"  Using 'fields' section with {len(fields)} keys")
                    elif isinstance(report, dict):
                        fields = report
                        debug(f"  Using entire report as fields with {len(fields)} keys")
                    else:
                        debug(f"  Report is not a dictionary, it's a {type(report).__name__}")
                        fields = {}
                    
                    # Add this file's data to the worksheet
                    structure_info = worksheets[safe_title]['structure_info']
                    self.add_data_row(
                        worksheet, 
                        next_row, 
                        file_name, 
                        fields, 
                        structure_info, 
                        max_list_lengths.get(safe_title, {}),
                        filter_text, 
                        apply_value_filters
                    )
                    
                    # Update the next row
                    worksheets[safe_title]['next_row'] = next_row + 1
            
            # Auto-adjust column widths for all worksheets
            debug("Adjusting column widths for all worksheets")
            for title, ws_info in worksheets.items():
                worksheet = ws_info['sheet']
                last_row = ws_info['next_row'] - 1
                column_count = ws_info['column_count']
                debug(f"  Adjusting widths for worksheet '{title}' with {column_count} columns and {last_row} rows")
                self.adjust_column_widths(worksheet, column_count + 1, last_row)  # +1 for safety
            
            debug(f"All processing complete. Processed {total_reports_processed} reports from {total_files} files.")
            debug(f"Created {len(worksheets)} worksheets.")
            debug(f"Saving workbook to {output_excel_path}")
            
            # Save the workbook
            workbook.save(output_excel_path)
            
            if callback:
                callback("status", f"Excel file created successfully at {output_excel_path}")
            return True
        
        except Exception as e:
            error_message = f"Error: {str(e)}"
            stack_trace = traceback.format_exc()
            
            if callback:
                callback("status", error_message)
                callback("debug", f"EXCEPTION: {error_message}")
                callback("debug", f"STACK TRACE: {stack_trace}")
            
            print(error_message)
            print(stack_trace)
            return False
    
    def get_column_count(self, structure_info):
        """Calculate the total number of columns needed based on structure info."""
        count = 1  # Start with 1 for the filename column
        
        for key in structure_info['keys']:
            nesting_depth = structure_info['nesting_depth'].get(key, 0)
            if nesting_depth > 1:
                count += nesting_depth
            else:
                count += 1
                
        return count
    
    def setup_headers(self, worksheet, structure_info):
        """Set up the headers for a worksheet."""
        # Set up the filename header
        filename_header = worksheet.cell(row=1, column=1, value="File Name")
        self.apply_cell_style(filename_header, self.header_style)
        
        # Add subtitle row if needed
        if structure_info['needs_subtitles']:
            subtitle_cell = worksheet.cell(row=2, column=1, value="")
            self.apply_cell_style(subtitle_cell, self.subtitle_style)
        
        # Set up field headers
        current_column = 2
        for key in sorted(structure_info['keys']):
            nesting_depth = structure_info['nesting_depth'].get(key, 0)
            
            # Set the header (key)
            header_cell = worksheet.cell(row=1, column=current_column, value=key)
            self.apply_cell_style(header_cell, self.header_style)
            
            if nesting_depth > 1:
                # This field has multiple items - needs subtitles
                # First, merge the header cell across all the items
                merge_end_column = current_column + nesting_depth - 1
                worksheet.merge_cells(
                    start_row=1, 
                    start_column=current_column, 
                    end_row=1, 
                    end_column=merge_end_column
                )
                
                # Center the merged header
                header_cell.alignment = Alignment(horizontal='center')
                
                # Create subtitle for each item
                for i in range(nesting_depth):
                    # Format subtitle: "Key - #1", "Key - #2", etc.
                    subtitle = f"{key} - #{i+1}"
                    
                    # Set the subtitle
                    subtitle_cell = worksheet.cell(row=2, column=current_column + i, value=subtitle)
                    self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                current_column += nesting_depth
            else:
                # This field has a single value or is not a list
                if structure_info['needs_subtitles']:
                    # If other fields have subtitles, add a blank subtitle cell for consistency
                    subtitle_cell = worksheet.cell(row=2, column=current_column, value="")
                    self.apply_cell_style(subtitle_cell, self.subtitle_style)
                
                current_column += 1
    
    def add_data_row(self, worksheet, row_num, file_name, fields, structure_info, max_list_lengths, 
                     filter_text="", apply_value_filters=True):
        """Add a row of data to the worksheet."""
        # Process filename to remove extension and filter text
        display_filename = JsonProcessor.process_filename(file_name, filter_text)
        
        # Write the processed filename
        worksheet.cell(row=row_num, column=1, value=display_filename)
        
        # Start with column 2 (after file name column)
        current_column = 2
        
        # Process each field
        for key in sorted(structure_info['keys']):
            value = fields.get(key, "")
            max_length = structure_info['nesting_depth'].get(key, 0)
            
            # Handle the different value types
            if max_length > 1:  # This field might have multiple items in some reports
                # Get the list value (or convert single value to a list)
                if isinstance(value, list):
                    list_value = value
                elif value or value == 0:  # Handle non-empty values including zero
                    list_value = [value]
                else:
                    list_value = []
                
                # Iterate through all possible positions, up to the maximum list length
                for i in range(max_length):
                    if i < len(list_value):
                        item = list_value[i]
                        # Filter and set the value
                        if apply_value_filters and isinstance(item, str):
                            # Apply text filtering to remove units
                            processed_item = TextFilter.remove_units(item)
                        else:
                            processed_item = item
                        
                        # Set the value
                        worksheet.cell(row=row_num, column=current_column, value=processed_item)
                    else:
                        # Skip this position - leave cell empty
                        worksheet.cell(row=row_num, column=current_column, value="")
                    
                    current_column += 1
            else:
                # This field has a single value or is not a list
                # Set the value (single value or first item of a list)
                if isinstance(value, list) and value:
                    value_to_set = value[0]
                else:
                    value_to_set = value
                
                # Apply text filtering if needed
                if apply_value_filters and isinstance(value_to_set, str):
                    value_to_set = TextFilter.remove_units(value_to_set)
                    
                # Set the processed value
                worksheet.cell(row=row_num, column=current_column, value=value_to_set)
                current_column += 1
    
    def adjust_column_widths(self, worksheet, num_columns, last_row):
        """Adjust column widths based on content."""
        for col_idx in range(1, num_columns + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            
            # Check header row(s) - always include these
            for row in range(1, min(3, last_row + 1)):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value:
                    text_length = len(str(cell.value))
                    max_length = max(max_length, text_length)
            
            # Check data rows
            for row in range(3, last_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value:
                    text_length = len(str(cell.value))
                    max_length = max(max_length, text_length)
            
            # Set column width (with some padding)
            if max_length > 0:
                adjusted_width = max_length + 2  # Add padding
                worksheet.column_dimensions[column].width = adjusted_width